import sys
import json
import PyPDF2
import pandas as pd
import re
from datetime import datetime
import io

def clean_description(text):
    """
    Pulisce e unifica la descrizione che può essere su più righe
    """
    # Rimuovi spazi multipli e caratteri speciali
    text = re.sub(r'\s+', ' ', text.strip())
    # Rimuovi caratteri non stampabili
    text = ''.join(char for char in text if char.isprintable() or char.isspace())
    return text.strip()

def extract_bank_data_from_pdf(pdf_content):
    """
    Estrae i dati bancari da un PDF di estratto conto
    Gestisce descrizioni su più righe
    """
    try:
        # Leggi il PDF
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_content))
        
        extracted_data = []
        
        # Pattern regex migliorati
        date_pattern = r'\b\d{2}[\/\-\.]\d{2}[\/\-\.]\d{4}\b'
        amount_pattern = r'[\d\.,]+(?:\s*€)?'
        
        for page_num, page in enumerate(pdf_reader.pages):
            text = page.extract_text()
            lines = text.split('\n')
            
            # Cerca l'inizio della tabella delle operazioni
            start_processing = False
            current_operation = None
            
            for i, line in enumerate(lines):
                line = line.strip()
                if not line:
                    continue
                
                # Identifica l'inizio della sezione operazioni
                if any(keyword in line.lower() for keyword in [
                    'data operazione', 'data valuta', 'descrizione', 
                    'data op.', 'data val.', 'causale', 'movimento'
                ]):
                    start_processing = True
                    continue
                
                if start_processing:
                    # Cerca pattern di data all'inizio della riga
                    date_matches = re.findall(date_pattern, line)
                    
                    # Se troviamo almeno una data, potrebbe essere una nuova operazione
                    if date_matches:
                        # Se abbiamo un'operazione in corso, salvala
                        if current_operation:
                            extracted_data.append(current_operation)
                        
                        # Inizia una nuova operazione
                        data_operazione = date_matches[0]
                        data_valuta = date_matches[1] if len(date_matches) > 1 else data_operazione
                        
                        # Trova gli importi nella riga
                        # Cerca numeri con formato italiano (virgola per decimali)
                        amounts = re.findall(r'\b\d{1,3}(?:\.\d{3})*(?:,\d{2})?\b', line)
                        
                        # Converti gli importi in float
                        numeric_amounts = []
                        for amt in amounts:
                            try:
                                # Converti formato italiano in float
                                amt_clean = amt.replace('.', '').replace(',', '.')
                                numeric_amounts.append(float(amt_clean))
                            except ValueError:
                                continue
                        
                        # Estrai la descrizione (tutto tra le date e gli importi)
                        descrizione_parts = []
                        
                        # Rimuovi le date dalla riga per estrarre la descrizione
                        temp_line = line
                        for date in date_matches:
                            temp_line = temp_line.replace(date, '', 1)
                        
                        # Rimuovi gli importi dalla riga
                        for amt in amounts:
                            temp_line = temp_line.replace(amt, '', 1)
                        
                        # Pulisci la descrizione
                        descrizione = clean_description(temp_line)
                        
                        # Determina accrediti e addebiti
                        accrediti = 0
                        addebiti = 0
                        
                        if numeric_amounts:
                            # Logica migliorata per determinare accrediti vs addebiti
                            # Cerca indicatori di addebito
                            is_debit = any(indicator in line.lower() for indicator in [
                                '-', 'addebito', 'prelievo', 'pagamento', 'commissione',
                                'canone', 'spese', 'imposta'
                            ])
                            
                            # Cerca indicatori di accredito
                            is_credit = any(indicator in line.lower() for indicator in [
                                '+', 'accredito', 'bonifico ricevuto', 'stipendio', 
                                'versamento', 'rimborso'
                            ])
                            
                            # Assegna l'importo
                            amount = numeric_amounts[-1]  # Prendi l'ultimo importo (di solito il principale)
                            
                            if is_debit and not is_credit:
                                addebiti = amount
                            elif is_credit and not is_debit:
                                accrediti = amount
                            else:
                                # Se non è chiaro, usa la posizione nella riga
                                # Questo dipende dal formato specifico della banca
                                if len(numeric_amounts) >= 2:
                                    accrediti = numeric_amounts[-2] if numeric_amounts[-2] > 0 else 0
                                    addebiti = numeric_amounts[-1] if numeric_amounts[-1] > 0 else 0
                                else:
                                    # Default: considera come accredito se positivo
                                    accrediti = amount if amount > 0 else 0
                        
                        current_operation = {
                            'dataOperazione': data_operazione,
                            'dataValuta': data_valuta,
                            'descrizione': descrizione,
                            'accrediti': accrediti,
                            'addebiti': addebiti
                        }
                    
                    elif current_operation and line:
                        # Questa potrebbe essere una continuazione della descrizione
                        # Aggiungi alla descrizione esistente
                        additional_desc = clean_description(line)
                        
                        # Verifica che non sia un'altra operazione o dati irrilevanti
                        if not re.search(r'\b\d{1,3}(?:\.\d{3})*(?:,\d{2})?\b', line) and len(additional_desc) > 3:
                            current_operation['descrizione'] += ' ' + additional_desc
                            current_operation['descrizione'] = clean_description(current_operation['descrizione'])
            
            # Aggiungi l'ultima operazione se presente
            if current_operation:
                extracted_data.append(current_operation)
        
        # Filtra operazioni valide (con almeno una data valida)
        valid_data = []
        for operation in extracted_data:
            if operation['dataOperazione'] and len(operation['descrizione']) > 2:
                valid_data.append(operation)
        
        return valid_data
        
    except Exception as e:
        raise Exception(f"Errore nell'estrazione dei dati: {str(e)}")

def convert_to_excel(data, output_path):
    """
    Converte i dati estratti in un file Excel con formattazione migliorata
    """
    try:
        df = pd.DataFrame(data)
        
        if df.empty:
            raise Exception("Nessun dato da convertire")
        
        # Rinomina le colonne per l'Excel
        df.columns = ['Data Operazione', 'Data Valuta', 'Descrizione', 'Accrediti', 'Addebiti']
        
        # Converti le date in formato datetime
        for date_col in ['Data Operazione', 'Data Valuta']:
            df[date_col] = pd.to_datetime(df[date_col], format='%d/%m/%Y', errors='coerce')
        
        # Pulisci le descrizioni
        df['Descrizione'] = df['Descrizione'].apply(lambda x: clean_description(str(x)) if pd.notna(x) else '')
        
        # Assicurati che gli importi siano numerici
        df['Accrediti'] = pd.to_numeric(df['Accrediti'], errors='coerce').fillna(0)
        df['Addebiti'] = pd.to_numeric(df['Addebiti'], errors='coerce').fillna(0)
        
        # Ordina per data operazione
        df = df.sort_values('Data Operazione')
        
        # Salva in Excel con formattazione
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Estratto Conto', index=False)
            
            # Formatta il foglio Excel
            worksheet = writer.sheets['Estratto Conto']
            
            # Imposta la larghezza delle colonne
            worksheet.column_dimensions['A'].width = 15  # Data Operazione
            worksheet.column_dimensions['B'].width = 15  # Data Valuta
            worksheet.column_dimensions['C'].width = 50  # Descrizione (più larga per testi lunghi)
            worksheet.column_dimensions['D'].width = 15  # Accrediti
            worksheet.column_dimensions['E'].width = 15  # Addebiti
            
            # Formatta le celle degli importi come valuta
            for row in range(2, len(df) + 2):
                if df.iloc[row-2]['Accrediti'] > 0:
                    worksheet[f'D{row}'].number_format = '€#,##0.00'
                if df.iloc[row-2]['Addebiti'] > 0:
                    worksheet[f'E{row}'].number_format = '€#,##0.00'
            
            # Formatta le date
            for row in range(2, len(df) + 2):
                worksheet[f'A{row}'].number_format = 'DD/MM/YYYY'
                worksheet[f'B{row}'].number_format = 'DD/MM/YYYY'
            
            # Aggiungi bordi e stile all'header
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Applica stile all'header
            for col in range(1, 6):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
        
        return True
        
    except Exception as e:
        raise Exception(f"Errore nella conversione Excel: {str(e)}")

def main():
    """
    Funzione principale per l'elaborazione da linea di comando
    """
    if len(sys.argv) != 3:
        print(json.dumps({"success": False, "error": "Uso: python pdf_processor.py <input_pdf> <output_excel>"}))
        sys.exit(1)
    
    input_pdf = sys.argv[1]
    output_excel = sys.argv[2]
    
    try:
        # Leggi il file PDF
        with open(input_pdf, 'rb') as file:
            pdf_content = file.read()
        
        # Estrai i dati
        data = extract_bank_data_from_pdf(pdf_content)
        
        if not data:
            print(json.dumps({
                "success": False, 
                "error": "Nessun dato bancario trovato nel PDF. Verifica che il file contenga un estratto conto valido."
            }))
            sys.exit(1)
        
        # Converti in Excel
        convert_to_excel(data, output_excel)
        
        # Restituisci il risultato
        result = {
            "success": True,
            "records_found": len(data),
            "output_file": output_excel,
            "preview": data[:5]  # Prime 5 righe per anteprima
        }
        
        print(json.dumps(result, ensure_ascii=False))
        
    except Exception as e:
        error_result = {
            "success": False,
            "error": str(e)
        }
        print(json.dumps(error_result, ensure_ascii=False))
        sys.exit(1)

if __name__ == "__main__":
    main()